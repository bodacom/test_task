import os
import pandas as pd
from tika import parser
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def frame_indexes(refs: list, dataset: str):
    '''
    Searches for the dataframe indexes based on the reference number
    Returns dictionary with reference number as a key and list of frame
    start and end indexes as a value
    '''
    indexes = dict()

    for ref, _ in enumerate(refs):
        if not ref == len(refs)-1:
            frame_start = dataset.find('P'+str(refs[ref])+'\n')
            frame_end = dataset.find('P'+str(refs[ref+1])+'\n')
        elif not refs[ref] == 102:
            frame_start = dataset.find('P'+str(refs[ref])+'\n')
            frame_end = dataset.find('P'+str(refs[ref]+1)+'\n')
        else:
            frame_start = dataset.find('P'+str(refs[ref])+'\n')
            frame_end = dataset.find('P'+str(104)+'\n')
        indexes[refs[ref]] = [frame_start, frame_end]
    return indexes


def topic_title(data_frame: str) -> tuple:
    '''
    Searches for the topic title in the data frame
    '''
    data_frame = data_frame.split('\n')[1:]
    title = ''
    
    for el in data_frame:
        if el.isupper():
            title = title + el
            elmnt = el
        else:
            break
    return title, elmnt


def presentation_abstract(data_frame: str) -> tuple:
    '''
    Searches for the abstract in the data frame
    Returns start position of the abstract based on the key word
    '''
    keywords = ['Introduction:',
                'Background:',
                'Introduction and Objectives:',
                'Background/Objective:',
                'Introduction/Objective:',
                ]
    position = 0
    found = False
    for keyword in keywords:
        if not data_frame.find(keyword) == -1:
            found, position = True, data_frame.find(keyword)
        if found:
            break
    
    return found, position


# opening pdf file
parsed_pdf = parser.from_file("file.pdf")

# saving content of pdf
data = parsed_pdf['content']

# generating reference numbers for abstracts
RANGE_START = 100
RANGE_END = 115

ref_nums = [i for i in range(RANGE_START, RANGE_END + 1) if not i == 103]

# generating frame indexes for separate abstracts
frames = frame_indexes(ref_nums, data)

parsed_data = []

for i, ref in enumerate(ref_nums):
    parsed_data.append([])
    # extracting data frame
    data_frm = data[frames[ref][0]:frames[ref][1]]
    
    # searching for the title and last fragment
    title, element = topic_title(data_frm)
    carrier1 = data_frm.find(element) + len(element)
    
    # searching for the abstract
    found, carrier_abstract = presentation_abstract(data_frm)
    if found:
        parsed_data[i].extend((data_frm[carrier1:carrier_abstract].strip(),
                              '',
                              '',
                              'P'+str(ref),
                              title,
                              data_frm[carrier_abstract:frames[ref][1]].strip()))
    else:
        parsed_data[i].extend(('Author not found',
                              '',
                              '',
                              'P'+str(ref),
                              title,
                              'Abstract not found'))
    
    for el in parsed_data[i]:
        print(el)

df = pd.DataFrame(parsed_data)
append_df_to_excel('Data Entry - 5th World Psoriasis & Psoriatic Arthritis Conference 2018 - Case format (2).xlsx',
                   df, header=None, index=False)
